#!/usr/bin/env python3
"""
DeepEP 16节点跨节点测试日志分析脚本
从rank_16.log中提取RDMA和NVL性能数据并生成Excel/CSV报告
"""

import re
import os
import csv
from typing import Dict, List, Tuple, Optional
import argparse

def parse_log_file(log_path: str) -> List[Dict]:
    """
    解析rank_16.log文件，提取跨节点性能数据
    
    Args:
        log_path: 日志文件路径
        
    Returns:
        包含测试数据的字典列表
    """
    if not os.path.exists(log_path):
        print(f"错误: 日志文件 {log_path} 不存在")
        return []
    
    with open(log_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    results = []
    
    # 使用正则表达式匹配每个配置块
    # 格式: [config] ... [tuning] Best dispatch ... [tuning] Best combine ...
    config_pattern = r'\[config\] num_tokens=(\d+), hidden=(\d+), num_topk=(\d+) ,num_experts=(\d+)'
    
    # 找到所有配置行
    config_matches = list(re.finditer(config_pattern, content))
    
    for i, config_match in enumerate(config_matches):
        # 提取配置信息
        num_tokens = int(config_match.group(1))
        hidden = int(config_match.group(2))
        num_topk = int(config_match.group(3))
        num_experts = int(config_match.group(4))
        
        # 确定当前配置块的范围
        start_pos = config_match.end()
        if i + 1 < len(config_matches):
            end_pos = config_matches[i + 1].start()
            section_content = content[start_pos:end_pos]
        else:
            section_content = content[start_pos:]
        
        # 提取dispatch性能数据
        # [tuning] Best dispatch (FP8): SMs 36, NVL chunk 8, RDMA chunk 8, transmit: 27.54 us, notify: 31.65 us, BW: 1.07 GB/s (RDMA), 2.95 GB/s (NVL)
        dispatch_pattern = r'\[tuning\] Best dispatch \(FP8\): SMs (\d+), NVL chunk (\d+), RDMA chunk (\d+), transmit: ([\d.]+) us, notify: ([\d.]+) us, BW: ([\d.]+) GB/s \(RDMA\), ([\d.]+) GB/s \(NVL\)'
        dispatch_match = re.search(dispatch_pattern, section_content)
        
        # 提取combine性能数据
        # [tuning] Best combine: SMs 36, NVL chunk 7, RDMA chunk 16, transmit: 55.25 us, notify: 30.98 us, BW: 1.04 GB/s (RDMA), 2.85 GB/s (NVL)
        combine_pattern = r'\[tuning\] Best combine: SMs (\d+), NVL chunk (\d+), RDMA chunk (\d+), transmit: ([\d.]+) us, notify: ([\d.]+) us, BW: ([\d.]+) GB/s \(RDMA\), ([\d.]+) GB/s \(NVL\)'
        combine_match = re.search(combine_pattern, section_content)
        
        # 构建结果字典
        result = {
            'num_tokens': num_tokens,
            'hidden': hidden,
            'num_topk': num_topk,
            'num_experts': num_experts,
        }
        
        # 添加dispatch数据
        if dispatch_match:
            result.update({
                'dispatch_sms': int(dispatch_match.group(1)),
                'dispatch_nvl_chunk': int(dispatch_match.group(2)),
                'dispatch_rdma_chunk': int(dispatch_match.group(3)),
                'dispatch_transmit_us': float(dispatch_match.group(4)),
                'dispatch_notify_us': float(dispatch_match.group(5)),
                'dispatch_rdma_bandwidth_gbps': float(dispatch_match.group(6)),
                'dispatch_nvl_bandwidth_gbps': float(dispatch_match.group(7)),
            })
        
        # 添加combine数据
        if combine_match:
            result.update({
                'combine_sms': int(combine_match.group(1)),
                'combine_nvl_chunk': int(combine_match.group(2)),
                'combine_rdma_chunk': int(combine_match.group(3)),
                'combine_transmit_us': float(combine_match.group(4)),
                'combine_notify_us': float(combine_match.group(5)),
                'combine_rdma_bandwidth_gbps': float(combine_match.group(6)),
                'combine_nvl_bandwidth_gbps': float(combine_match.group(7)),
            })
        
        
        results.append(result)
    
    return results

def parse_ll_log_file(log_path: str) -> List[Dict]:
    """
    解析 *ll.log 文件，提取两类行各生成一行：
    1) return_recv_hook=True 的 send/recv time 行
    2) return_recv_hook=False 的 bandwidth/avg_t 行
    每匹配一行即返回一条记录。
    """
    if not os.path.exists(log_path):
        print(f"错误: 日志文件 {log_path} 不存在")
        return []

    results: List[Dict] = []

    # 预编译正则
    alloc_pattern = re.compile(r"Allocating buffer size: ([\d.]+) MB ...")
    timing_pattern = re.compile(
        r"\[rank \d+\] num_tokens=(\d+), hidden=(\d+), num_experts=(\d+), num_topk=(\d+), return_recv_hook=True "
        r"Dispatch send/recv time: ([\d.]+) \+ ([\d.]+) us \| Combine send/recv time: ([\d.]+) \+ ([\d.]+) us"
    )
    bw_pattern = re.compile(
        r"\[rank \d+\] num_tokens=(\d+), hidden=(\d+), num_experts=(\d+), num_topk=(\d+), return_recv_hook=False "
        r"Dispatch bandwidth: ([\d.]+) GB/s, avg_t=([\d.]+) us \| Combine bandwidth: ([\d.]+) GB/s, avg_t=([\d.]+) us"
    )

    current_alloc_mb: Optional[float] = None

    with open(log_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            # 捕获最近的 buffer 大小
            alloc_m = alloc_pattern.search(line)
            if alloc_m:
                try:
                    current_alloc_mb = float(alloc_m.group(1))
                except ValueError:
                    current_alloc_mb = None
                continue

            # 匹配 timing 行（return_recv_hook=True）
            m1 = timing_pattern.search(line)
            if m1:
                try:
                    num_tokens = int(m1.group(1))
                    hidden = int(m1.group(2))
                    num_experts = int(m1.group(3))
                    num_topk = int(m1.group(4))
                    dispatch_send = float(m1.group(5))
                    dispatch_recv = float(m1.group(6))
                    combine_send = float(m1.group(7))
                    combine_recv = float(m1.group(8))
                except ValueError:
                    continue

                row: Dict = {
                    'num_tokens': num_tokens,
                    'hidden': hidden,
                    'num_topk': num_topk,
                    'num_experts': num_experts,
                    'return_recv_hook': True,
                    'dispatch_transmit_us': dispatch_send,
                    'dispatch_notify_us': dispatch_recv,
                    'combine_transmit_us': combine_send,
                    'combine_notify_us': combine_recv,
                }
                if current_alloc_mb is not None:
                    row['data_size_mb'] = current_alloc_mb
                results.append(row)
                continue

            # 匹配 bandwidth 行（return_recv_hook=False）
            m2 = bw_pattern.search(line)
            if m2:
                try:
                    num_tokens = int(m2.group(1))
                    hidden = int(m2.group(2))
                    num_experts = int(m2.group(3))
                    num_topk = int(m2.group(4))
                    dispatch_bw = float(m2.group(5))
                    dispatch_avg_t = float(m2.group(6))
                    combine_bw = float(m2.group(7))
                    combine_avg_t = float(m2.group(8))
                except ValueError:
                    continue

                row2: Dict = {
                    'num_tokens': num_tokens,
                    'hidden': hidden,
                    'num_topk': num_topk,
                    'num_experts': num_experts,
                    'return_recv_hook': False,
                    'dispatch_bandwidth_gbps': dispatch_bw,
                    'dispatch_avg_t_us': dispatch_avg_t,
                    'combine_bandwidth_gbps': combine_bw,
                    'combine_avg_t_us': combine_avg_t,
                }
                if current_alloc_mb is not None:
                    row2['data_size_mb'] = current_alloc_mb
                results.append(row2)
                continue

    return results

def collect_log_files(log_dir: str) -> List[str]:
    """
    仅收集指定目录下(不递归)的 .log 文件
    """
    if not os.path.isdir(log_dir):
        return []
    log_files: List[str] = []
    for name in os.listdir(log_dir):
        path = os.path.join(log_dir, name)
        if os.path.isfile(path) and name.endswith('.log'):
            log_files.append(path)
    return sorted(log_files)

def _sanitize_sheet_name(name: str) -> str:
    """
    清理Excel工作表名称（去除非法字符并截断至31字符）。
    """
    invalid = set(':\\/?*[]')
    clean = ''.join('_' if ch in invalid else ch for ch in name)
    if not clean:
        clean = 'sheet'
    return clean[:31]

def create_csv_report(all_data: List[Dict], output_path: str):
    """
    创建CSV报告
    
    Args:
        all_data: 所有测试数据
        output_path: 输出CSV文件路径
    """
    if not all_data:
        print("警告: 没有找到有效的测试数据")
        return
    
    # 定义重要列的顺序
    priority_columns = [
        'num_tokens', 'hidden', 'num_topk', 'num_experts',
        'data_size_mb', 'total_time_us', 'total_throughput_gbps',
        
        # Dispatch 指标
        'dispatch_sms', 'dispatch_nvl_chunk', 'dispatch_rdma_chunk',
        'dispatch_transmit_us', 'dispatch_notify_us',
        'dispatch_rdma_bandwidth_gbps', 'dispatch_nvl_bandwidth_gbps',
        
        # Combine 指标
        'combine_sms', 'combine_nvl_chunk', 'combine_rdma_chunk',
        'combine_transmit_us', 'combine_notify_us',
        'combine_rdma_bandwidth_gbps', 'combine_nvl_bandwidth_gbps',
        
        # 汇总指标
        'avg_rdma_bandwidth_gbps', 'avg_nvl_bandwidth_gbps',
        'total_transmit_time_us', 'total_notify_time_us',
    ]
    
    # 收集所有可能的列名
    all_columns = set()
    for data in all_data:
        all_columns.update(data.keys())
    
    # 排序列名
    sorted_columns = []
    for col in priority_columns:
        if col in all_columns:
            sorted_columns.append(col)
            all_columns.remove(col)
    
    # 添加剩余的列
    sorted_columns.extend(sorted(all_columns))
    
    # 写入CSV文件
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=sorted_columns)
        writer.writeheader()
        
        # 按num_tokens排序后写入数据
        sorted_data = sorted(all_data, key=lambda x: x.get('num_tokens', 0))
        for row in sorted_data:
            # 将数值四舍五入到4位小数
            rounded_row = {}
            for k, v in row.items():
                if isinstance(v, float):
                    rounded_row[k] = round(v, 4)
                else:
                    rounded_row[k] = v
            writer.writerow(rounded_row)
    
    print(f"✅ CSV报告已生成: {output_path}")

def create_summary_report(all_data: List[Dict], output_dir: str):
    """
    创建汇总报告
    """
    if not all_data:
        return
    
    summary_path = os.path.join(output_dir, "internode_summary.txt")
    
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("DeepEP 16节点跨节点测试性能汇总\n")
        f.write("=" * 50 + "\n\n")
        
        # 基本信息
        f.write(f"总测试配置: {len(all_data)}\n")
        
        # Token数范围
        token_values = [d.get('num_tokens', 0) for d in all_data]
        f.write(f"Token数范围: {min(token_values)} - {max(token_values)}\n")
        f.write(f"专家数: {all_data[0].get('num_experts', 'N/A')}\n")
        f.write(f"隐藏层大小: {all_data[0].get('hidden', 'N/A')}\n\n")
        
        # 性能指标统计
        metrics = [
            ('dispatch_rdma_bandwidth_gbps', 'Dispatch RDMA带宽', 'GB/s'),
            ('dispatch_nvl_bandwidth_gbps', 'Dispatch NVL带宽', 'GB/s'),
            ('combine_rdma_bandwidth_gbps', 'Combine RDMA带宽', 'GB/s'),
            ('combine_nvl_bandwidth_gbps', 'Combine NVL带宽', 'GB/s'),
            ('avg_rdma_bandwidth_gbps', '平均RDMA带宽', 'GB/s'),
            ('avg_nvl_bandwidth_gbps', '平均NVL带宽', 'GB/s'),
            ('total_throughput_gbps', '整体吞吐率', 'GB/s'),
            ('dispatch_transmit_us', 'Dispatch传输时间', 'μs'),
            ('combine_transmit_us', 'Combine传输时间', 'μs'),
            ('total_time_us', '总时间', 'μs'),
        ]
        
        f.write("性能指标统计:\n")
        f.write("-" * 30 + "\n")
        
        for metric_key, metric_name, unit in metrics:
            values = [d.get(metric_key, 0) for d in all_data if metric_key in d]
            if values:
                f.write(f"{metric_name}:\n")
                f.write(f"  平均: {sum(values)/len(values):.4f} {unit}\n")
                f.write(f"  最大: {max(values):.4f} {unit}\n")
                f.write(f"  最小: {min(values):.4f} {unit}\n")
                f.write("\n")
        
        # 配置优化信息
        f.write("最优配置信息:\n")
        f.write("-" * 30 + "\n")
        
        for data in all_data:
            tokens = data.get('num_tokens', 0)
            f.write(f"Token={tokens}:\n")
            f.write(f"  Dispatch: SMs={data.get('dispatch_sms', 'N/A')}, "
                   f"NVL chunk={data.get('dispatch_nvl_chunk', 'N/A')}, "
                   f"RDMA chunk={data.get('dispatch_rdma_chunk', 'N/A')}\n")
            f.write(f"  Combine: SMs={data.get('combine_sms', 'N/A')}, "
                   f"NVL chunk={data.get('combine_nvl_chunk', 'N/A')}, "
                   f"RDMA chunk={data.get('combine_rdma_chunk', 'N/A')}\n")
            if 'total_throughput_gbps' in data:
                f.write(f"  整体吞吐率: {data['total_throughput_gbps']:.2f} GB/s\n")
            f.write("\n")
    
    print(f"✅ 汇总报告已生成: {summary_path}")

def try_create_excel_report(all_data: List[Dict], output_path: str):
    """
    尝试创建Excel报告（如果有openpyxl的话）
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("❌ 未安装openpyxl，无法生成Excel。请安装后重试: pip install openpyxl")
        return False

    # 组装列头（优先重要列）
    priority_columns = [
        'num_tokens', 'hidden', 'num_topk', 'num_experts',
        'data_size_mb', 'total_time_us', 'total_throughput_gbps',
        'dispatch_sms', 'dispatch_nvl_chunk', 'dispatch_rdma_chunk',
        'dispatch_transmit_us', 'dispatch_notify_us',
        'dispatch_rdma_bandwidth_gbps', 'dispatch_nvl_bandwidth_gbps',
        'combine_sms', 'combine_nvl_chunk', 'combine_rdma_chunk',
        'combine_transmit_us', 'combine_notify_us',
        'combine_rdma_bandwidth_gbps', 'combine_nvl_bandwidth_gbps',
        'avg_rdma_bandwidth_gbps', 'avg_nvl_bandwidth_gbps',
        'total_transmit_time_us', 'total_notify_time_us',
    ]
    all_columns = set()
    for data in all_data:
        all_columns.update(data.keys())
    sorted_columns: List[str] = []
    for col in priority_columns:
        if col in all_columns:
            sorted_columns.append(col)
            all_columns.remove(col)
    sorted_columns.extend(sorted(all_columns))

    wb = Workbook()
    ws = wb.active
    ws.title = '原始数据'
    ws.append(sorted_columns)

    sorted_data = sorted(all_data, key=lambda x: x.get('num_tokens', 0))
    for row in sorted_data:
        excel_row = []
        for col in sorted_columns:
            value = row.get(col, '')
            if isinstance(value, float):
                value = round(value, 4)
            excel_row.append(value)
        ws.append(excel_row)

    output_dir = os.path.dirname(os.path.abspath(output_path)) or '.'
    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel报告已生成: {output_path}")
    return True

def try_create_excel_report_multi_sheet(logfile_to_data: Dict[str, List[Dict]], output_path: str) -> bool:
    """
    生成多sheet的Excel：每个日志文件对应一个sheet，sheet名为日志文件名。
    优先使用pandas；若不可用，退回openpyxl。
    """
    # 过滤空数据
    logfile_to_data = {k: v for k, v in logfile_to_data.items() if v}
    if not logfile_to_data:
        print("警告: 没有找到有效的测试数据")
        return False

    # 尝试pandas
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        print("❌ 未安装openpyxl，无法生成Excel。请安装后重试: pip install openpyxl")
        return False

    wb = Workbook()
    # 默认工作表将用于第一个sheet
    first = True
    used_names = set()
    for path, rows in logfile_to_data.items():
        base = os.path.basename(path)
        name = _sanitize_sheet_name(base)
        original = name
        idx = 1
        while name in used_names:
            suffix = f"_{idx}"
            name = _sanitize_sheet_name((original[:31 - len(suffix)]) + suffix)
            idx += 1
        used_names.add(name)

        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)

        # 计算列集合，优先重要列
        priority_columns = [
            'num_tokens', 'hidden', 'num_topk', 'num_experts',
            'data_size_mb', 'total_time_us', 'total_throughput_gbps',
            'dispatch_sms', 'dispatch_nvl_chunk', 'dispatch_rdma_chunk',
            'dispatch_transmit_us', 'dispatch_notify_us',
            'dispatch_rdma_bandwidth_gbps', 'dispatch_nvl_bandwidth_gbps',
            'combine_sms', 'combine_nvl_chunk', 'combine_rdma_chunk',
            'combine_transmit_us', 'combine_notify_us',
            'combine_rdma_bandwidth_gbps', 'combine_nvl_bandwidth_gbps',
            'avg_rdma_bandwidth_gbps', 'avg_nvl_bandwidth_gbps',
            'total_transmit_time_us', 'total_notify_time_us',
        ]
        all_columns = set()
        for r in rows:
            all_columns.update(r.keys())
        ordered = []
        for c in priority_columns:
            if c in all_columns:
                ordered.append(c)
                all_columns.remove(c)
        ordered.extend(sorted(all_columns))

        # 表头
        ws.append(ordered)

        # 按num_tokens排序写入
        rows_sorted = sorted(rows, key=lambda x: x.get('num_tokens', 0))
        for r in rows_sorted:
            row_vals = []
            for c in ordered:
                v = r.get(c, '')
                if isinstance(v, float):
                    v = round(v, 4)
                row_vals.append(v)
            ws.append(row_vals)

    output_dir = os.path.dirname(os.path.abspath(output_path)) or '.'
    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel报告已生成: {output_path}")
    return True

def main():
    parser = argparse.ArgumentParser(description='从目录下所有.log生成多Sheet的Excel报告')
    parser.add_argument('--log-dir', default='/home/xutingz/workspace/bench/deepep/v1',
                       help='包含.log文件的目录路径 (默认: /home/xutingz/workspace/bench/deepep)')
    parser.add_argument('--output-excel', default='./internode_performance.xlsx',
                       help='输出Excel文件路径 (默认: ./internode_performance.xlsx)')
    args = parser.parse_args()
    
    print("🔍 开始解析DeepEP 16节点跨节点测试日志...")
    print(f"日志目录: {args.log_dir}")
    print(f"Excel输出: {args.output_excel}")
    print("=" * 50)
    
    # 收集日志文件
    log_files = collect_log_files(args.log_dir)
    if not log_files:
        print("❌ 没有找到任何 .log 日志文件")
        return
    print(f"找到 {len(log_files)} 个日志文件，开始解析...")

    logfile_to_data: Dict[str, List[Dict]] = {}
    for log_file in log_files:
        print(f"正在解析: {os.path.basename(log_file)}...")
        if log_file.endswith('ll.log'):
            data = parse_ll_log_file(log_file)
        else:
            data = parse_log_file(log_file)
        if data:
            logfile_to_data[log_file] = data

    if not logfile_to_data:
        print("❌ 没有提取到任何有效的测试数据")
        return

    # 生成多sheet Excel
    print("\n📊 尝试生成多Sheet Excel报告...")
    excel_success = try_create_excel_report_multi_sheet(logfile_to_data, args.output_excel)
    
    # 汇总打印（跨所有文件）
    all_data: List[Dict] = []
    for rows in logfile_to_data.values():
        all_data.extend(rows)
    if all_data:
        print(f"\n📈 数据汇总:")
        print(f"  - Token数范围: {min(d.get('num_tokens', 0) for d in all_data)} - {max(d.get('num_tokens', 0) for d in all_data)}")
        if 'num_experts' in all_data[0]:
            print(f"  - 专家数: {all_data[0].get('num_experts', 'N/A')}")
        avg_rdma_bw = [d.get('avg_rdma_bandwidth_gbps', 0) for d in all_data if 'avg_rdma_bandwidth_gbps' in d]
        avg_nvl_bw = [d.get('avg_nvl_bandwidth_gbps', 0) for d in all_data if 'avg_nvl_bandwidth_gbps' in d]
        if avg_rdma_bw:
            print(f"  - 平均RDMA带宽: {sum(avg_rdma_bw)/len(avg_rdma_bw):.2f} GB/s")
        if avg_nvl_bw:
            print(f"  - 平均NVL带宽: {sum(avg_nvl_bw)/len(avg_nvl_bw):.2f} GB/s")
        total_throughput = [d.get('total_throughput_gbps', 0) for d in all_data if 'total_throughput_gbps' in d]
        if total_throughput:
            print(f"  - 整体平均吞吐率: {sum(total_throughput)/len(total_throughput):.2f} GB/s")
    
    print(f"\n🎉 报告生成完成!")
    if excel_success:
        print(f"  - Excel: {args.output_excel}")

if __name__ == '__main__':
    main()
